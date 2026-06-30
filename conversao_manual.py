from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Erro: biblioteca 'openpyxl' não encontrada. Instale com: pip install openpyxl",
        file=sys.stderr,
    )
    raise

DEFAULT_OUTPUT_DIR = Path(
    r"F:\04 - Operacional\Operacional\SUPORTE OPERACIONAL\Prêmios\Export Prêmios"
)
DEFAULT_SOURCE_FILE = Path(
    r"F:\04 - Operacional\Operacional\SUPORTE OPERACIONAL\Prêmios\Controle_Premios V9.6 nova_x.xlsm"
)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def _normalize_header(value: Any) -> str:
    text = _normalize_text(value)
    return re.sub(r"[^A-Z0-9]+", "", text)


def _excel_date_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _map_verb(tipo_pagamento: Any, observacao: Any) -> str:
    tipo = _normalize_text(tipo_pagamento)

    if tipo == "EXTRA":
        return "Hora Extra"
    if tipo == "DIARIA":
        return "Diária"
    if tipo == "PROMOCAO":
        return "Promoção"

    if tipo == "PREMIO":
        obs = _normalize_text(observacao)
        if "TESTE" in obs:
            return "Teste (Descreva)"
        if "COBERTURA" in obs:
            return "Cobertura (Descreva)"
        if "REMOCAO" in obs:
            return "Remoção"
        if "VIDRO" in obs:
            return "Limpeza de Vidros"
        return "Outros"

    return "Outros"


def _remove_monetary_tokens(text: str) -> str:
    cleaned = text
    # Remove formatos monetários comuns: R$ 1.234,56 | 1234,56 | 1234.56
    patterns = [
        r"R\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{2})?",
        r"R\$\s*\d+(?:,\d{2})?",
        r"\b\d{1,3}(?:\.\d{3})+,\d{2}\b",
        r"\b\d+,\d{2}\b",
        r"\b\d+\.\d{2}\b",
    ]
    for pattern in patterns:
        cleaned = re.sub(pattern, " ", cleaned, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", cleaned).strip(" -;,:")


def padronizar_observacao(texto: Any) -> str:
    if texto is None:
        return ""
    original = str(texto).strip()
    if not original:
        return ""

    texto_sem_valor = _remove_monetary_tokens(original)
    normalized = _normalize_text(texto_sem_valor)

    # 1) Cobertura de cargo
    if "COBERTURA" in normalized:
        if "ENCARREGADO" in normalized:
            return "COBERTURA DE ENCARREGADO"
        return "COBERTURA DE LÍDER"

    # 2) Teste de cargo
    if "TESTE" in normalized:
        if "ENCARREGADO" in normalized:
            return "TESTE DE ENCARREGADO"
        if "LIDER" in normalized:
            return "TESTE DE LÍDER"

    # 3) Prêmio de viagem / implantação
    if "VIAGEM" in normalized or "IMPLANTACAO" in normalized:
        return "PRÊMIO DE VIAGEM"

    # 4) Prêmio de apoio
    if "APOIO" in normalized:
        return "PRÊMIO DE APOIO"

    # 5) Adicional de função
    if "COPEIRA" in normalized:
        return "ADICIONAL DE COPEIRA"
    if "CAFE" in normalized:
        return "ADICIONAL DE CAFÉ"
    if "CAMERA FRIA" in normalized or "CAMARA FRIA" in normalized:
        return "ADICIONAL DE CÂMARA FRIA"
    if "LIMPEZA DE VIDRO" in normalized or "LIMPEZA DE VIDROS" in normalized:
        return "ADICIONAL DE LIMPEZA DE VIDROS"

    # 6) Bom desempenho
    if "BOM DESEMPENHO" in normalized:
        return "PRÊMIO DE BOM DESEMPENHO"

    # 7) Auxílio
    if "AUXILIO" in normalized:
        return "PRÊMIO POR AUXÍLIO"

    # 8) Pagamento junto com hora extra
    if "JUNTO COM AS HORAS EXTRAS" in normalized or "HORA EXTRA DO DIA" in normalized:
        return "PRÊMIO REFERENTE A TRABALHO EM DIA ESPECÍFICO"

    # 9) Não identificado: mantém texto original (sem valores monetários)
    return texto_sem_valor

def _determinar_roteiro(observacao: Any) -> str:
    """
    Determina o roteiro com base na observação.
    Se contiver 'VEX' -> retorna 'VEX'
    Caso contrário -> retorna 'FOLHA'
    """
    if observacao is None or observacao == "":
        return "FOLHA"
    
    texto_normalizado = _normalize_text(observacao)
    
    if "VEX" in texto_normalizado:
        return "VEX"
    
    return "FOLHA"

def _headers_map(header_row: Iterable[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, header in enumerate(header_row, start=1):
        key = _normalize_header(header)
        if key:
            mapping[key] = idx
    return mapping


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _period_yyyymm(value: Any) -> str:
    dt = _to_date(value)
    if not dt:
        return ""
    return f"{dt.year:04d}{dt.month:02d}"


def _periodo_pt_br(year: int, month: int) -> str:
    meses = {
        1: "Jan",
        2: "Fev",
        3: "Mar",
        4: "Abr",
        5: "Mai",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Set",
        10: "Out",
        11: "Nov",
        12: "Dez",
    }
    if month not in meses:
        return ""
    return f"{meses[month]}_{year:04d}"


def _format_reward_value(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"

    text = str(value).strip()
    if not text:
        return ""

    # Aceita formatos como 1.234,56 | 1234,56 | 1234.56
    normalized = text.replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")

    try:
        return f"{float(normalized):.2f}"
    except ValueError:
        # Se não for numérico, mantém como está.
        return text


def convert(
    source_file: Path,
    target_file: Path,
    source_sheet: str = "Base_Pagamentos",
    target_sheet: str | None = None,
    output_file: Path | None = None,
    filter_year: int | None = None,
    filter_month: int | None = None,
) -> int:
    src_wb = load_workbook(source_file, data_only=True)
    if source_sheet not in src_wb.sheetnames:
        raise ValueError(
            f"Aba '{source_sheet}' não encontrada em {source_file}. "
            f"Abas disponíveis: {src_wb.sheetnames}"
        )
    src_ws = src_wb[source_sheet]

    tgt_wb = load_workbook(target_file)
    if target_sheet is None:
        tgt_ws = tgt_wb[tgt_wb.sheetnames[0]]
    else:
        if target_sheet not in tgt_wb.sheetnames:
            raise ValueError(
                f"Aba '{target_sheet}' não encontrada em {target_file}. "
                f"Abas disponíveis: {tgt_wb.sheetnames}"
            )
        tgt_ws = tgt_wb[target_sheet]

    src_headers = _headers_map(c.value for c in src_ws[1])
    tgt_headers = _headers_map(c.value for c in tgt_ws[1])

    order_type_key = "ORDERTYPE"
    order_type_col = tgt_headers.get(order_type_key)
    if order_type_col is None:
        order_type_col = tgt_ws.max_column + 1
        tgt_ws.cell(row=1, column=order_type_col, value="order_type")
        tgt_headers[order_type_key] = order_type_col

    observacao_key = "OBSERVACAO"
    observacao_col = tgt_headers.get(observacao_key)
    if observacao_col is None:
        observacao_col = tgt_ws.max_column + 1
        tgt_ws.cell(row=1, column=observacao_col, value="Observação")
        tgt_headers[observacao_key] = observacao_col
        
    roteiro_key = "ROTEIRO"
    roteiro_col = tgt_headers.get(roteiro_key)
    if roteiro_col is None:
        roteiro_col = tgt_ws.max_column + 1
        tgt_ws.cell(row=1, column=roteiro_col, value="Roteiro")
        tgt_headers[roteiro_key] = roteiro_col

    required_source = {
        "RE": "RE",
        "FAVORECIDO": "Favorecido",
        "TIPOPAGAMENTO": "Tipo_Pagamento",
        "VALOR": "Valor",
        "DATATRABALHADA": "Data_Trabalhada",
    }
    missing_src = [label for key, label in required_source.items() if key not in src_headers]
    if missing_src:
        raise ValueError(
            "Colunas obrigatórias não encontradas na planilha origem: "
            + ", ".join(missing_src)
        )

    required_target = ["STATUS", "EMPLOYEEID", "EMPLOYEENAME", "VERBNAME", "REWARDVALUE", "PERIOD"]
    missing_tgt = [k for k in required_target if k not in tgt_headers]
    if missing_tgt:
        raise ValueError(
            "Colunas obrigatórias não encontradas na planilha destino: "
            + ", ".join(missing_tgt)
        )

    idx_re = src_headers["RE"]
    idx_favorecido = src_headers["FAVORECIDO"]
    idx_tipo = src_headers["TIPOPAGAMENTO"]
    idx_valor = src_headers["VALOR"]
    idx_data = src_headers["DATATRABALHADA"]
    idx_obs = src_headers.get("OBSERVACAO")

    out_status = tgt_headers["STATUS"]
    out_employee_id = tgt_headers["EMPLOYEEID"]
    out_employee_name = tgt_headers["EMPLOYEENAME"]
    out_verb_name = tgt_headers["VERBNAME"]
    out_reward_value = tgt_headers["REWARDVALUE"]
    out_period = tgt_headers["PERIOD"]
    out_order_type = tgt_headers["ORDERTYPE"]
    out_observacao = tgt_headers["OBSERVACAO"]
    out_roteiro = tgt_headers["ROTEIRO"]

    existing_last_row = tgt_ws.max_row
    target_data_cols = [
        out_status,
        out_employee_id,
        out_employee_name,
        out_verb_name,
        out_reward_value,
        out_period,
    ]
    for row_idx in range(2, existing_last_row + 1):
        if any(tgt_ws.cell(row=row_idx, column=col).value not in (None, "") for col in target_data_cols):
            tgt_ws.cell(row=row_idx, column=out_order_type, value="SISTEMA")
            
            # Preencher ROTEIRO para linhas existentes também
            obs_existente = tgt_ws.cell(row=row_idx, column=out_observacao).value
            roteiro_valor = _determinar_roteiro(obs_existente)
            tgt_ws.cell(row=row_idx, column=out_roteiro, value=roteiro_valor)

    inserted = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        re_value = row[idx_re - 1]
        nome_value = row[idx_favorecido - 1]
        tipo_value = row[idx_tipo - 1]
        valor_value = row[idx_valor - 1]
        data_value = row[idx_data - 1]
        obs_original = row[idx_obs - 1] if idx_obs else None
        obs_value = padronizar_observacao(obs_original)

        if all(v in (None, "") for v in (re_value, nome_value, tipo_value, valor_value, data_value, obs_value)):
            continue

        worked_date = _to_date(data_value)
        if filter_year is not None and filter_month is not None:
            if worked_date is None:
                continue
            if worked_date.year != filter_year or worked_date.month != filter_month:
                continue

        period_value = _period_yyyymm(data_value)
        verb_name = _map_verb(tipo_value, obs_value)
        reward_value = _format_reward_value(valor_value)
        if verb_name == "Promoção":
            reward_value = "0.00"

        new_row = tgt_ws.max_row + 1
        tgt_ws.cell(row=new_row, column=out_status, value="PAGO")
        tgt_ws.cell(row=new_row, column=out_employee_id, value=re_value)
        tgt_ws.cell(row=new_row, column=out_employee_name, value=nome_value)
        tgt_ws.cell(row=new_row, column=out_verb_name, value=verb_name)
        tgt_ws.cell(row=new_row, column=out_reward_value, value=reward_value)
        tgt_ws.cell(row=new_row, column=out_period, value=period_value)
        tgt_ws.cell(row=new_row, column=out_order_type, value="MANUAL")
        tgt_ws.cell(row=new_row, column=out_observacao, value=obs_value)
        inserted += 1
        tgt_ws.cell(row=new_row, column=out_roteiro, value=_determinar_roteiro(obs_original))

    output_path = output_file or target_file
    tgt_wb.save(output_path)
    return inserted


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Converte dados da aba Base_Pagamentos (Controle_Premios) e adiciona no relatorio_premios.xlsx"
        )
    )
    parser.add_argument(
        "--source",
        default=str(DEFAULT_SOURCE_FILE),
        help="Caminho do arquivo origem (.xlsm)",
    )
    parser.add_argument(
        "--target",
        default=None,
        help="Caminho do arquivo destino (.xlsx)",
    )
    parser.add_argument(
        "--source-sheet",
        default="Base_Pagamentos",
        help="Nome da aba origem",
    )
    parser.add_argument(
        "--target-sheet",
        default=None,
        help="Nome da aba destino (padrão: primeira aba)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Arquivo de saída. Se omitido, sobrescreve o arquivo destino.",
    )
    parser.add_argument(
        "--periodo",
        default=None,
        help="Filtro por período no formato YYYYMM (ex.: 202602).",
    )
    parser.add_argument(
        "--ano",
        type=int,
        default=None,
        help="Ano do filtro de Data_Trabalhada (usar com --mes).",
    )
    parser.add_argument(
        "--mes",
        type=int,
        default=None,
        help="Mês do filtro de Data_Trabalhada (1-12, usar com --ano).",
    )
    parser.add_argument(
        "--visual",
        action="store_true",
        help="Abre janelas para escolher arquivos e período.",
    )
    return parser


def _selecionar_arquivos_e_periodo_visual() -> tuple[Path, Path, Path | None, int | None, int | None]:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("Conversão de Lançamentos Manuais")
    root.geometry("860x320")
    root.resizable(False, False)

    source_var = tk.StringVar(value=str(DEFAULT_SOURCE_FILE))
    target_var = tk.StringVar()
    out_dir_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))
    periodo_var = tk.StringVar()
    result: Dict[str, Any] = {}

    frame = ttk.Frame(root, padding=14)
    frame.pack(fill="both", expand=True)

    ttk.Label(
        frame,
        text=(
            "Selecione os arquivos e o período do processamento.\n"
            "Período opcional no formato YYYYMM (ex.: 202601)."
        ),
    ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 12))

    def selecionar_source() -> None:
        path = filedialog.askopenfilename(
            title="Selecione a planilha de lançamentos manuais",
            filetypes=[("Excel", "*.xlsm *.xlsx *.xls"), ("Todos os arquivos", "*.*")],
        )
        if path:
            source_var.set(path)

    def selecionar_target() -> None:
        path = filedialog.askopenfilename(
            title="Selecione a planilha relatorio_premios",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos os arquivos", "*.*")],
        )
        if path:
            target_var.set(path)

    def selecionar_saida() -> None:
        path = filedialog.askdirectory(title="Selecione a pasta para salvar o arquivo gerado")
        if path:
            out_dir_var.set(path)

    ttk.Label(frame, text="1) Planilha de lançamentos manuais:").grid(row=1, column=0, sticky="w")
    ttk.Entry(frame, textvariable=source_var, width=88).grid(row=2, column=0, sticky="we", padx=(0, 8))
    ttk.Button(frame, text="Selecionar", command=selecionar_source).grid(row=2, column=1, sticky="w")

    ttk.Label(frame, text="2) Planilha relatorio_premios:").grid(row=3, column=0, sticky="w", pady=(10, 0))
    ttk.Entry(frame, textvariable=target_var, width=88).grid(row=4, column=0, sticky="we", padx=(0, 8))
    ttk.Button(frame, text="Selecionar", command=selecionar_target).grid(row=4, column=1, sticky="w")

    ttk.Label(frame, text="3) Pasta de saída do novo arquivo:").grid(row=5, column=0, sticky="w", pady=(10, 0))
    ttk.Entry(frame, textvariable=out_dir_var, width=88).grid(row=6, column=0, sticky="we", padx=(0, 8))
    ttk.Button(frame, text="Selecionar", command=selecionar_saida).grid(row=6, column=1, sticky="w")

    ttk.Label(frame, text="4) Período para filtro (YYYYMM, opcional):").grid(row=7, column=0, sticky="w", pady=(10, 0))
    ttk.Entry(frame, textvariable=periodo_var, width=20).grid(row=8, column=0, sticky="w")

    def confirmar() -> None:
        source_txt = source_var.get().strip()
        target_txt = target_var.get().strip()
        out_dir_txt = out_dir_var.get().strip()
        periodo_txt = periodo_var.get().strip()

        if not source_txt:
            messagebox.showerror("Validação", "Selecione a planilha de lançamentos manuais.")
            return
        if not target_txt:
            messagebox.showerror("Validação", "Selecione a planilha relatorio_premios.")
            return
        if not out_dir_txt:
            messagebox.showerror("Validação", "Selecione a pasta de saída.")
            return

        filter_year = None
        filter_month = None
        if periodo_txt:
            if not re.fullmatch(r"\d{6}", periodo_txt):
                messagebox.showerror("Validação", "Período inválido. Use YYYYMM, por exemplo 202601.")
                return
            filter_year = int(periodo_txt[:4])
            filter_month = int(periodo_txt[4:6])
            if not (1 <= filter_month <= 12):
                messagebox.showerror("Validação", "Mês inválido no período informado.")
                return

        nome_arquivo = "relatorio_premios_convertido.xlsx"
        if filter_year is not None and filter_month is not None:
            nome_arquivo = f"{_periodo_pt_br(filter_year, filter_month)}.xlsx"

        result["source"] = Path(source_txt)
        result["target"] = Path(target_txt)
        result["output"] = Path(out_dir_txt) / nome_arquivo
        result["filter_year"] = filter_year
        result["filter_month"] = filter_month
        root.destroy()

    ttk.Button(frame, text="Gerar Arquivo", command=confirmar).grid(row=9, column=0, sticky="w", pady=(16, 0))
    ttk.Button(frame, text="Cancelar", command=root.destroy).grid(row=9, column=1, sticky="w", pady=(16, 0))

    root.mainloop()

    if not result:
        raise ValueError("Operação cancelada.")
    return (
        result["source"],
        result["target"],
        result["output"],
        result["filter_year"],
        result["filter_month"],
    )


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    source: Path
    target: Path
    output: Path | None
    filter_year = None
    filter_month = None

    run_visual = args.visual or len(sys.argv) == 1
    if run_visual:
        source, target, output_visual, fy_visual, fm_visual = _selecionar_arquivos_e_periodo_visual()
        output = output_visual
        filter_year = fy_visual
        filter_month = fm_visual
    else:
        if not args.target:
            raise ValueError("No modo não visual, informe --target.")
        source = Path(args.source) if args.source else DEFAULT_SOURCE_FILE
        target = Path(args.target)
        output = Path(args.output) if args.output else None

        if args.periodo:
            periodo_txt = str(args.periodo).strip()
            if not re.fullmatch(r"\d{6}", periodo_txt):
                raise ValueError("O parâmetro --periodo deve estar no formato YYYYMM, exemplo: 202602")
            filter_year = int(periodo_txt[:4])
            filter_month = int(periodo_txt[4:6])
        elif args.ano is not None or args.mes is not None:
            if args.ano is None or args.mes is None:
                raise ValueError("Informe os dois parâmetros juntos: --ano e --mes")
            filter_year = args.ano
            filter_month = args.mes

    if filter_month is not None and not (1 <= filter_month <= 12):
        raise ValueError("O mês deve estar entre 1 e 12.")
    if not source.exists():
        raise FileNotFoundError(f"Arquivo de origem não encontrado: {source}")
    if not target.exists():
        raise FileNotFoundError(f"Arquivo de destino não encontrado: {target}")

    if output is None:
        if filter_year is not None and filter_month is not None:
            output = DEFAULT_OUTPUT_DIR / f"{_periodo_pt_br(filter_year, filter_month)}.xlsx"
        else:
            output = DEFAULT_OUTPUT_DIR / "relatorio_premios_convertido.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    inserted = convert(
        source_file=source,
        target_file=target,
        source_sheet=args.source_sheet,
        target_sheet=args.target_sheet,
        output_file=output,
        filter_year=filter_year,
        filter_month=filter_month,
    )
    dest = output if output else target
    print(f"Concluído. Linhas adicionadas: {inserted}. Arquivo salvo em: {dest}")


if __name__ == "__main__":
    main()
