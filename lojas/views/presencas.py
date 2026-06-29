import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import re
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from usuarios.permissions import IsGestaoOrAdministrador
from unidecode import unidecode
from ..models import Loja

def normalizar_nome(valor):
    """
    Normaliza strings removendo acentos, convertendo para maiúsculo e removendo espaços.
    """
    if valor is None:
        return ""
    return unidecode(str(valor)).strip().upper()

def clean_re(valor):
    """
    Por que existe: Extrai apenas os dígitos numéricos do RE do colaborador para
    garantir que possamos cruzar o mesmo código independentemente de prefixos ou zeros à esquerda.
    """
    if valor is None:
        return ""
    digits = "".join(re.findall(r"\d+", str(valor)))
    return str(int(digits)) if digits else ""

def parse_date_str(valor):
    """
    Por que existe: Extrai a data e retorna um objeto datetime para que possamos
    calcular a diferença de dias de forma correta e encontrar a loja mais próxima temporalmente.
    """
    if valor is None:
        return None
    s = str(valor).strip()
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", s)
    if match:
        d, m, y = match.groups()
        try:
            return datetime(int(y), int(m), int(d))
        except Exception:
            return None
    return None

def tem_valor(valor):
    """
    Por que existe: Identifica se uma célula possui dados válidos de batida (não nulo e não texto 'None'/'Null').
    """
    if valor is None:
        return False
    s = str(valor).strip()
    if s == "" or s.lower() == "none" or s.lower() == "null":
        return False
    return True

@api_view(["POST"])
@permission_classes([IsAuthenticated, IsGestaoOrAdministrador])
def importar_presencas_api(request):
    """
    Por que existe: Esta view processa o upload de duas planilhas (Punch Report e Controle de Ponto).
    Ela calcula a quantidade de presenças (com base em ter ao menos batida de Entrada)
    e folgas (contendo Folga/DOMINGO/Descanso/Feriado) por colaborador e dia.
    Para cada evento, a loja associada é buscada no Punch Report como a loja mais próxima temporalmente do dia analisado.
    No final, cruza com as lojas do banco de dados para gerar o consolidado de presenças e folgas por loja.
    """
    punch_file = request.FILES.get("punch_file")
    controle_file = request.FILES.get("controle_file")
    
    if not punch_file or not controle_file:
        return Response(
            {"error": "Ambos os arquivos (Punch Report e Controle de Ponto) são obrigatórios."},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    try:
        # 1. Carrega a planilha de Controle de Ponto
        wb_ctrl = load_workbook(filename=controle_file, read_only=True)
        sheet_ctrl = wb_ctrl.active
        
        controle_data = defaultdict(dict) # { re: { date_obj: (has_presence, is_folga) } }
        
        row_count = 0
        for row in sheet_ctrl.iter_rows(values_only=True):
            row_count += 1
            if row_count <= 2: # Pula cabeçalhos (linhas 1 e 2)
                continue
                
            if len(row) < 20: # Garante que a linha possui colunas suficientes (até T / 19)
                continue
                
            sobrenomes = row[0]   # Coluna A
            data_raw = row[4]     # Coluna E
            permissao = row[5]    # Coluna F
            entrou = row[7]       # Coluna H
            saiu = row[19]        # Coluna T
            
            re_val = clean_re(sobrenomes)
            dt = parse_date_str(data_raw)
            
            if not re_val or not dt:
                continue
                
            # Regra de presença: Entrada (H) e Saída (T)
            has_presence = tem_valor(entrou) and tem_valor(saiu)
            
            # Regra de Folga: contiver Folga, DOMINGO, Descanso ou Feriado (ignora caixa alta/baixa)
            is_folga = False
            if permissao:
                p_lower = str(permissao).strip().lower()
                if any(w in p_lower for w in ["folga", "domingo", "descanso", "feriado"]):
                    is_folga = True
                    
            controle_data[re_val][dt] = (has_presence, is_folga)
            
        # 2. Carrega a planilha de Punch Report
        wb_punch = load_workbook(filename=punch_file, read_only=True)
        if "Con Marcas" not in wb_punch.sheetnames:
            return Response(
                {"error": "A planilha de Punch Report não contém a aba 'Con Marcas'."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        sheet_punch = wb_punch["Con Marcas"]
        
        punches_by_employee = defaultdict(list) # { re: [ (date_obj, marcacion) ] }
        
        row_count = 0
        for row in sheet_punch.iter_rows(values_only=True):
            row_count += 1
            if row_count <= 3: # Pula cabeçalhos (linhas 1, 2 e 3)
                continue
                
            if len(row) < 10: # Garante que tem a coluna Marcación (J / 9)
                continue
                
            apellidos = row[0]    # Coluna A
            fecha = row[4]        # Coluna E
            marcacion = row[9]    # Coluna J
            
            re_val = clean_re(apellidos)
            dt = parse_date_str(fecha)
            
            if not re_val or not dt:
                continue
                
            punches_by_employee[re_val].append((dt, marcacion))
            
        # Ordena cronologicamente os punches de cada funcionário
        for re_val in punches_by_employee:
            punches_by_employee[re_val].sort(key=lambda x: x[0])
            
        # Função para achar a loja correta daquele momento
        def obter_loja_do_momento(re_colaborador, data_evento):
            emp_punches = punches_by_employee.get(re_colaborador)
            if not emp_punches:
                return "Grupo Não Informado"
                
            closest_store = None
            min_diff = None
            for p_dt, store in emp_punches:
                diff = abs((p_dt - data_evento).total_seconds())
                if min_diff is None or diff < min_diff:
                    min_diff = diff
                    closest_store = store
            return closest_store if closest_store else "Grupo Não Informado"
            
        # 3. Mapeamento das Lojas do banco de dados para busca rápida
        mapa_lojas = {}
        # nome_totvs
        for l in Loja.objects.exclude(nome_totvs=""):
            n = normalizar_nome(l.nome_totvs)
            if n:
                mapa_lojas[n] = l
        # nome_gestao
        for l in Loja.objects.exclude(nome_gestao=""):
            n = normalizar_nome(l.nome_gestao)
            if n:
                mapa_lojas[n] = l
        # nome_referencia
        for l in Loja.objects.all():
            n = normalizar_nome(l.nome_referencia)
            if n:
                mapa_lojas[n] = l
        # nome_geovictoria
        for l in Loja.objects.exclude(nome_geovictoria=""):
            n = normalizar_nome(l.nome_geovictoria)
            if n:
                mapa_lojas[n] = l
                
        # 4. Processamento cruzado por loja
        # { loja: { 'presencas': int, 'folgas': int, 'colaboradores': set } }
        lojas_stats = defaultdict(lambda: {
            "presencas": 0,
            "folgas": 0,
            "colaboradores": set()
        })
        
        total_colaboradores_unicos = set()
        
        for re_val, dates_dict in controle_data.items():
            for dt, (has_presence, is_folga) in dates_dict.items():
                if not (has_presence or is_folga):
                    continue
                    
                # Acha a loja no momento do evento
                store_name = obter_loja_do_momento(re_val, dt)
                store_name = str(store_name).strip() if store_name else "Grupo Não Informado"
                
                stats = lojas_stats[store_name]
                stats["colaboradores"].add(re_val)
                total_colaboradores_unicos.add(re_val)
                
                if has_presence:
                    stats["presencas"] += 1
                elif is_folga:
                    stats["folgas"] += 1
                    
        # 5. Consolidação final dos resultados por loja
        linhas_relatorio = []
        total_presencas = 0
        total_folgas = 0
        lojas_encontradas = 0
        lojas_nao_encontradas = 0
        grupos_nao_encontrados_set = set()
        
        for grupo, stats in lojas_stats.items():
            grupo_norm = normalizar_nome(grupo)
            loja = mapa_lojas.get(grupo_norm)
            
            unique_employees = len(stats["colaboradores"])
            total_presencas += stats["presencas"]
            total_folgas += stats["folgas"]
            
            if loja:
                lojas_encontradas += 1
                status_loja = "encontrada"
                loja_id = loja.id
                loja_ref = loja.nome_referencia
                loja_geo = loja.nome_geovictoria
                cc = loja.centro_de_custo
            else:
                lojas_nao_encontradas += 1
                status_loja = "nao_encontrada"
                loja_id = None
                loja_ref = "Não encontrada no banco"
                loja_geo = ""
                cc = ""
                if grupo != "Grupo Não Informado":
                    grupos_nao_encontrados_set.add(grupo)
                    
            linhas_relatorio.append({
                "grupo_planilha": grupo,
                "loja_id": loja_id,
                "loja_referencia": loja_ref,
                "loja_geovictoria": loja_geo,
                "centro_de_custo": cc,
                "presencas": stats["presencas"],
                "folgas": stats["folgas"],
                "funcionarios_unicos": unique_employees,
                "status": status_loja,
            })
            
        # Ordena as divergentes no topo, depois por presenças decrescente
        linhas_relatorio.sort(key=lambda x: (0 if x["status"] == "nao_encontrada" else 1, -x["presencas"]))
        
        resultado = {
            "summary": {
                "total_presencas": total_presencas,
                "total_folgas": total_folgas,
                "total_lojas_encontradas": lojas_encontradas,
                "total_lojas_nao_encontradas": lojas_nao_encontradas,
                "colaboradores_unicos": len(total_colaboradores_unicos),
            },
            "unmatched_groups": sorted(list(grupos_nao_encontrados_set)),
            "rows": linhas_relatorio
        }
        
        return Response(resultado, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({"error": f"Erro interno ao processar planilhas: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
